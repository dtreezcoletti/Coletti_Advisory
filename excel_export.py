"""
Coletti OS v2.5.5 — Excel Export Module
Case No. 24D-1003 | Coletti v. Brown

Produces a formatted .xlsx forensic evidence package from a ColettiOS instance.
"""

from __future__ import annotations

import io
from datetime import datetime
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from coletti_os import ColettiOS


# ---------------------------------------------------------------------------
# Shared style constants
# ---------------------------------------------------------------------------

_FILL_HEADER = PatternFill("solid", fgColor="0D1117")        # dark navy
_FILL_SUBTOTAL = PatternFill("solid", fgColor="1A3A5C")      # medium navy
_FILL_GRAND_TOTAL = PatternFill("solid", fgColor="C0392B")   # deep red
_FILL_DISSIPATION = PatternFill("solid", fgColor="FFD7D7")   # light red
_FILL_ALT_ROW = PatternFill("solid", fgColor="F2F2F2")       # light grey
_FILL_INCOME_FRAUD_HIGHLIGHT = PatternFill("solid", fgColor="FF4444")  # bright red
_FILL_COVER_ACCENT = PatternFill("solid", fgColor="0D1117")

_FONT_HEADER = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_FONT_SUBTOTAL = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_FONT_GRAND_TOTAL = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
_FONT_BODY = Font(name="Calibri", size=11)
_FONT_BOLD = Font(name="Calibri", bold=True, size=11)
_FONT_DISSIPATION = Font(name="Calibri", bold=True, size=11, color="8B0000")
_FONT_TITLE = Font(name="Calibri", bold=True, size=20, color="0D1117")
_FONT_SUBTITLE = Font(name="Calibri", bold=True, size=13, color="1A3A5C")
_FONT_CONFIDENTIAL = Font(name="Calibri", bold=True, size=14, color="C0392B")

_FMT_AMOUNT = '"$"#,##0.00'
_FMT_DATE = "YYYY-MM-DD"
_FMT_PCT = '0.00"%"'

_THIN = Side(border_style="thin", color="CCCCCC")
_THIN_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_BOTTOM_BORDER = Border(bottom=Side(border_style="medium", color="0D1117"))


# ---------------------------------------------------------------------------
# Helper utilities
# ---------------------------------------------------------------------------

def _apply_header_row(ws, row_idx: int, values: list, widths: list = None):
    """Write a header row with dark navy fill and white bold text."""
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.fill = _FILL_HEADER
        cell.font = _FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER


def _apply_subtotal_row(ws, row_idx: int, values: list):
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.fill = _FILL_SUBTOTAL
        cell.font = _FONT_SUBTOTAL
        cell.alignment = Alignment(horizontal="right" if col > 1 else "left")
        cell.border = _THIN_BORDER


def _apply_grand_total_row(ws, row_idx: int, values: list):
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.fill = _FILL_GRAND_TOTAL
        cell.font = _FONT_GRAND_TOTAL
        cell.alignment = Alignment(horizontal="right" if col > 1 else "left")
        cell.border = _THIN_BORDER


def _set_col_widths(ws, widths: list):
    """Set column widths from a list of (col_idx, width) tuples."""
    for col_idx, width in widths:
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _write_kv_block(ws, start_row: int, start_col: int, pairs: list, label_width: int = 28):
    """Write a series of label: value pairs as two-column block."""
    for i, (label, value) in enumerate(pairs):
        r = start_row + i
        lc = ws.cell(row=r, column=start_col, value=label)
        lc.font = _FONT_BOLD
        lc.alignment = Alignment(horizontal="left")
        vc = ws.cell(row=r, column=start_col + 1, value=value)
        vc.font = _FONT_BODY
        vc.alignment = Alignment(horizontal="left")
    return start_row + len(pairs)


# ---------------------------------------------------------------------------
# ExcelExporter
# ---------------------------------------------------------------------------

class ExcelExporter:
    """
    Produces a formatted .xlsx forensic evidence package from a ColettiOS instance.

    Usage::

        data = ExcelExporter().export(sys_instance)
        with open("report.xlsx", "wb") as f:
            f.write(data)
    """

    def export(self, sys_instance: "ColettiOS") -> bytes:
        """
        Generate and return the workbook as raw bytes.

        Parameters
        ----------
        sys_instance:
            A fully initialised ColettiOS object.
        """
        wb = Workbook()
        # Remove the default empty sheet
        wb.remove(wb.active)

        self._build_cover(wb, sys_instance)
        self._build_transaction_ledger(wb, sys_instance)
        self._build_dissipation_analysis(wb, sys_instance)
        self._build_income_fraud(wb, sys_instance)
        self._build_case_valuation(wb, sys_instance)
        self._build_forensic_summary(wb, sys_instance)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ------------------------------------------------------------------
    # Sheet 1: Cover
    # ------------------------------------------------------------------

    def _build_cover(self, wb: Workbook, sys: "ColettiOS"):
        ws = wb.create_sheet("Cover")
        ws.sheet_view.showGridLines = False

        # Column widths
        ws.column_dimensions["A"].width = 4
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 42
        ws.column_dimensions["D"].width = 4

        # Row heights
        ws.row_dimensions[1].height = 8
        ws.row_dimensions[2].height = 50
        ws.row_dimensions[3].height = 24
        ws.row_dimensions[4].height = 20

        # ── Title banner (row 2) ──────────────────────────────────────
        ws.merge_cells("B2:C2")
        title_cell = ws["B2"]
        title_cell.value = "FORENSIC FINANCIAL EVIDENCE PACKAGE"
        title_cell.font = _FONT_TITLE
        title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # ── Confidential stamp (row 3) ────────────────────────────────
        ws.merge_cells("B3:C3")
        conf_cell = ws["B3"]
        conf_cell.value = "CONFIDENTIAL — ATTORNEY WORK PRODUCT"
        conf_cell.font = _FONT_CONFIDENTIAL
        conf_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Separator row
        for col in ["B", "C"]:
            ws[f"{col}4"].fill = _FILL_HEADER

        ws.row_dimensions[5].height = 12

        # ── Case metadata ─────────────────────────────────────────────
        metadata = [
            ("Firm Name:",          sys.enterprise.firm_name),
            ("Analyst:",            sys.enterprise.founder),
            ("Case Number:",        sys.litigation.case_number),
            ("Case Name:",          "Coletti v. Brown"),
            ("Court:",              sys.litigation.jurisdiction),
            ("Judge:",              sys.litigation.judge),
            ("Generation Date:",    datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("ColettiOS Version:",  ColettiOS_VERSION_FROM_SYS(sys)),
        ]
        current_row = 6
        for label, value in metadata:
            ws.row_dimensions[current_row].height = 18
            lc = ws.cell(row=current_row, column=2, value=label)
            lc.font = _FONT_BOLD
            lc.alignment = Alignment(horizontal="right", vertical="center")
            vc = ws.cell(row=current_row, column=3, value=value)
            vc.font = _FONT_BODY
            vc.alignment = Alignment(horizontal="left", vertical="center")
            current_row += 1

        current_row += 1

        # ── Active subpoenas ──────────────────────────────────────────
        ws.row_dimensions[current_row].height = 18
        hdr = ws.cell(row=current_row, column=2, value="Active Subpoenas:")
        hdr.font = _FONT_BOLD
        current_row += 1
        for sub in sys.litigation.active_subpoenas:
            ws.row_dimensions[current_row].height = 16
            sc = ws.cell(row=current_row, column=3, value=f"• {sub}")
            sc.font = _FONT_BODY
            current_row += 1

        current_row += 1

        # ── Case dates summary ────────────────────────────────────────
        cv = sys.case_valuation
        ws.row_dimensions[current_row].height = 18
        ws.cell(row=current_row, column=2, value="Key Case Dates:").font = _FONT_BOLD
        current_row += 1
        date_pairs = [
            ("Marriage Start:", cv.case_dates.marriage_start),
            ("Separation Date:", cv.case_dates.separation_date),
            ("Filing Date:", cv.case_dates.filing_date),
            ("Marriage Duration:", f"{cv.case_dates.marriage_years():.2f} years"),
        ]
        for label, value in date_pairs:
            ws.row_dimensions[current_row].height = 16
            lc = ws.cell(row=current_row, column=2, value=label)
            lc.font = Font(name="Calibri", size=11)
            lc.alignment = Alignment(horizontal="right")
            vc = ws.cell(row=current_row, column=3, value=value)
            vc.font = _FONT_BODY
            current_row += 1

    # ------------------------------------------------------------------
    # Sheet 2: Transaction Ledger
    # ------------------------------------------------------------------

    def _build_transaction_ledger(self, wb: Workbook, sys: "ColettiOS"):
        ws = wb.create_sheet("Transaction Ledger")

        headers = [
            "Date", "Amount", "Description", "Category",
            "Dissipation Flag", "Cumulative Dissipation",
        ]
        _apply_header_row(ws, 1, headers)
        ws.freeze_panes = "A2"

        # Column widths
        _set_col_widths(ws, [
            (1, 14), (2, 14), (3, 40), (4, 22), (5, 18), (6, 22),
        ])

        transactions = sys.forensics.transactions
        cumulative_dissipation = 0.0
        total_amount = 0.0
        total_dissipation = 0.0

        for i, tx in enumerate(transactions):
            row_idx = i + 2
            is_dis = tx.is_marital_dissipation
            if is_dis:
                cumulative_dissipation += tx.amount

            total_amount += tx.amount
            if is_dis:
                total_dissipation += tx.amount

            fill = _FILL_DISSIPATION if is_dis else (
                _FILL_ALT_ROW if i % 2 == 1 else PatternFill()
            )
            font_amount = _FONT_DISSIPATION if is_dis else _FONT_BODY

            cells = [
                (1, tx.effective_date),
                (2, tx.amount),
                (3, tx.description),
                (4, tx.category),
                (5, "YES" if is_dis else ""),
                (6, cumulative_dissipation if is_dis else None),
            ]
            for col, val in cells:
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.fill = fill
                cell.border = _THIN_BORDER
                if col == 2:
                    cell.font = font_amount
                    cell.number_format = _FMT_AMOUNT
                    cell.alignment = Alignment(horizontal="right")
                elif col == 6 and val is not None:
                    cell.number_format = _FMT_AMOUNT
                    cell.alignment = Alignment(horizontal="right")
                    cell.font = font_amount if is_dis else _FONT_BODY
                else:
                    cell.font = _FONT_BODY
                    if col == 1:
                        cell.number_format = _FMT_DATE

        # Total row
        total_row = len(transactions) + 2
        if not transactions:
            total_row = 3

        total_values = [
            "TOTAL", total_amount, "", "",
            f"{total_dissipation:,.2f} dissipated", total_dissipation,
        ]
        _apply_grand_total_row(ws, total_row, total_values)
        for col in [2, 6]:
            ws.cell(row=total_row, column=col).number_format = _FMT_AMOUNT
            ws.cell(row=total_row, column=col).alignment = Alignment(horizontal="right")

    # ------------------------------------------------------------------
    # Sheet 3: Dissipation Analysis
    # ------------------------------------------------------------------

    def _build_dissipation_analysis(self, wb: Workbook, sys: "ColettiOS"):
        ws = wb.create_sheet("Dissipation Analysis")
        ws.sheet_view.showGridLines = False

        _set_col_widths(ws, [
            (1, 36), (2, 18), (3, 18), (4, 18),
        ])

        transactions = sys.forensics.transactions
        total_amount = sum(t.amount for t in transactions)
        dissipation_amount = sum(t.amount for t in transactions if t.is_marital_dissipation)
        dissipation_rate = (dissipation_amount / total_amount * 100) if total_amount else 0.0

        # ── Summary table ─────────────────────────────────────────────
        ws.row_dimensions[1].height = 20
        title = ws.cell(row=1, column=1, value="DISSIPATION ANALYSIS SUMMARY")
        title.font = _FONT_SUBTITLE

        _apply_header_row(ws, 2, ["Metric", "Value"])

        summary_rows = [
            ("Total Transactions", len(transactions)),
            ("Total Funds Reviewed", total_amount),
            ("Dissipation Amount", dissipation_amount),
            ("Dissipation Rate %", dissipation_rate / 100),
        ]
        for i, (label, val) in enumerate(summary_rows):
            r = i + 3
            lc = ws.cell(row=r, column=1, value=label)
            lc.font = _FONT_BODY
            lc.border = _THIN_BORDER
            vc = ws.cell(row=r, column=2, value=val)
            vc.border = _THIN_BORDER
            if "Amount" in label or "Reviewed" in label:
                vc.number_format = _FMT_AMOUNT
                vc.alignment = Alignment(horizontal="right")
            elif "Rate" in label:
                vc.number_format = '0.00%'
                vc.alignment = Alignment(horizontal="right")
            else:
                vc.font = _FONT_BODY

        # ── Category breakdown ────────────────────────────────────────
        cat_start = 9
        ws.row_dimensions[cat_start - 1].height = 8
        cat_title = ws.cell(row=cat_start - 1, column=1, value="CATEGORY BREAKDOWN")
        cat_title.font = _FONT_SUBTITLE

        _apply_header_row(ws, cat_start, ["Category", "Total Amount", "Dissipation Amount", "Txn Count"])

        # Aggregate by category
        categories: dict = {}
        for tx in transactions:
            cat = tx.category or "Uncategorized"
            if cat not in categories:
                categories[cat] = {"total": 0.0, "dissipation": 0.0, "count": 0}
            categories[cat]["total"] += tx.amount
            categories[cat]["count"] += 1
            if tx.is_marital_dissipation:
                categories[cat]["dissipation"] += tx.amount

        # Add placeholder if empty
        if not categories:
            categories["No Data"] = {"total": 0.0, "dissipation": 0.0, "count": 0}

        chart_data_start = cat_start + 1
        for i, (cat_name, stats) in enumerate(sorted(categories.items())):
            r = cat_start + 1 + i
            ws.cell(row=r, column=1, value=cat_name).border = _THIN_BORDER
            amt_cell = ws.cell(row=r, column=2, value=stats["total"])
            amt_cell.number_format = _FMT_AMOUNT
            amt_cell.alignment = Alignment(horizontal="right")
            amt_cell.border = _THIN_BORDER
            dis_cell = ws.cell(row=r, column=3, value=stats["dissipation"])
            dis_cell.number_format = _FMT_AMOUNT
            dis_cell.alignment = Alignment(horizontal="right")
            dis_cell.border = _THIN_BORDER
            cnt_cell = ws.cell(row=r, column=4, value=stats["count"])
            cnt_cell.alignment = Alignment(horizontal="center")
            cnt_cell.border = _THIN_BORDER

        chart_data_end = cat_start + len(categories)

        # ── Bar chart ─────────────────────────────────────────────────
        try:
            chart = BarChart()
            chart.type = "col"
            chart.title = "Amount by Category"
            chart.y_axis.title = "Amount ($)"
            chart.x_axis.title = "Category"
            chart.style = 10
            chart.width = 20
            chart.height = 12

            data_ref = Reference(
                ws,
                min_col=2,
                max_col=3,
                min_row=cat_start,
                max_row=chart_data_end,
            )
            cats_ref = Reference(
                ws,
                min_col=1,
                min_row=chart_data_start,
                max_row=chart_data_end,
            )
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            ws.add_chart(chart, f"A{chart_data_end + 3}")
        except Exception:
            pass  # Chart is optional; never block export

    # ------------------------------------------------------------------
    # Sheet 4: Income Fraud
    # ------------------------------------------------------------------

    def _build_income_fraud(self, wb: Workbook, sys: "ColettiOS"):
        ws = wb.create_sheet("Income Fraud")
        ws.sheet_view.showGridLines = False

        _set_col_widths(ws, [
            (1, 36), (2, 20),
        ])

        fraud = sys.case_valuation.income_fraud

        ws.row_dimensions[1].height = 22
        title = ws.cell(row=1, column=1, value="SWORN vs. VERIFIED INCOME COMPARISON")
        title.font = _FONT_SUBTITLE

        _apply_header_row(ws, 2, ["Income Category", "Amount"])

        rows = [
            ("Sworn Annual Income (Affidavit)", fraud.sworn_annual, False),
            ("Verified W-2 Income (Dreamliner)", fraud.verified_w2, False),
            ("Verified 1099 Income (R.E. Garrison)", fraud.verified_1099, False),
            ("Verified Total Income", fraud.verified_total, False),
            ("Concealment Amount", fraud.concealment_amount, True),
            ("Concealment %", fraud.concealment_pct / 100, True),
        ]

        for i, (label, val, highlight) in enumerate(rows):
            r = i + 3
            lc = ws.cell(row=r, column=1, value=label)
            vc = ws.cell(row=r, column=2, value=val)

            if highlight:
                lc.fill = PatternFill("solid", fgColor="FF4444")
                vc.fill = PatternFill("solid", fgColor="FF4444")
                lc.font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
                vc.font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
            else:
                alt = _FILL_ALT_ROW if i % 2 == 1 else PatternFill()
                lc.fill = alt
                vc.fill = alt
                lc.font = _FONT_BODY
                vc.font = _FONT_BODY

            lc.border = _THIN_BORDER
            vc.border = _THIN_BORDER
            vc.alignment = Alignment(horizontal="right")

            if "%" in label:
                vc.number_format = '0.00%'
            else:
                vc.number_format = _FMT_AMOUNT

        # Add notes block
        note_row = len(rows) + 5
        ws.cell(row=note_row, column=1, value="SOURCE DOCUMENTATION").font = _FONT_BOLD
        note_row += 1
        notes = [
            "• W-2: Dreamliner Aircraft Services — Paystubs obtained 2026-02-08",
            "• 1099: R.E. Garrison Trucking Company — 2024 Form 1099-NEC obtained 2026-02-15",
            "• Sworn: Financial Affidavit filed 2025-05-27",
        ]
        for note in notes:
            ws.cell(row=note_row, column=1, value=note).font = Font(name="Calibri", size=10, italic=True)
            note_row += 1

    # ------------------------------------------------------------------
    # Sheet 5: Case Valuation
    # ------------------------------------------------------------------

    def _build_case_valuation(self, wb: Workbook, sys: "ColettiOS"):
        ws = wb.create_sheet("Case Valuation")
        ws.sheet_view.showGridLines = False

        _set_col_widths(ws, [
            (1, 42), (2, 18),
        ])

        cv = sys.case_valuation
        row = 1

        # ── Title ─────────────────────────────────────────────────────
        ws.row_dimensions[row].height = 24
        t = ws.cell(row=row, column=1, value="CASE VALUATION — THREE-TIER DAMAGE ANALYSIS")
        t.font = _FONT_SUBTITLE
        row += 2

        # ── Tier 1 ────────────────────────────────────────────────────
        row = self._valuation_tier(
            ws, row,
            tier_label="TIER 1: MOTION RELIEF (IMMEDIATE)",
            items=[
                ("Suit Money", cv.tier1.suit_money),
                ("Income Concealment — Proven", cv.tier1.income_concealment_proven),
                ("Income Concealment — Suspected", cv.tier1.income_concealment_suspected),
                ("Income Concealment Total", cv.tier1.income_concealment_total),
                ("Animal Equalization", cv.tier1.animal_equalization),
                ("Pendente Lite Arrearage", cv.tier1.pendente_lite_arrearage),
                ("Pendente Lite Monthly (reference)", cv.tier1.pendente_lite_monthly),
                ("King Personal Sanctions", cv.tier1.king_personal_sanctions),
            ],
            subtotal=cv.tier1.subtotal,
        )
        row += 1

        # ── Tier 2 ────────────────────────────────────────────────────
        row = self._valuation_tier(
            ws, row,
            tier_label="TIER 2: TRIAL DAMAGES",
            items=[
                ("Homemaker Contributions", cv.tier2.homemaker_contributions),
                ("Human Capital Loss", cv.tier2.human_capital_loss),
                ("Business Sabotage", cv.tier2.business_sabotage),
                ("Property Division", cv.tier2.property_division),
                ("Alimony in Solido", cv.tier2.alimony_in_solido),
                ("Marital Fault Damages", cv.tier2.marital_fault_damages),
            ],
            subtotal=cv.tier2.subtotal,
        )
        row += 1

        # ── Tier 3 ────────────────────────────────────────────────────
        row = self._valuation_tier(
            ws, row,
            tier_label="TIER 3: PUNITIVE DAMAGES",
            items=[
                ("Assault Punitive", cv.tier3.assault_punitive),
                ("Economic Destruction Punitive", cv.tier3.economic_destruction_punitive),
            ],
            subtotal=cv.tier3.total,
        )
        row += 1

        # ── Grand total ───────────────────────────────────────────────
        ws.row_dimensions[row].height = 22
        gt = ws.cell(row=row, column=1, value="GRAND TOTAL (ALL TIERS CAPPED)")
        gt.fill = _FILL_GRAND_TOTAL
        gt.font = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
        gt.border = _THIN_BORDER
        gv = ws.cell(row=row, column=2, value=cv.total_capped)
        gv.fill = _FILL_GRAND_TOTAL
        gv.font = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
        gv.number_format = _FMT_AMOUNT
        gv.alignment = Alignment(horizontal="right")
        gv.border = _THIN_BORDER

    def _valuation_tier(
        self, ws, start_row: int, tier_label: str,
        items: list, subtotal: float
    ) -> int:
        """Write one tier section. Returns the next available row."""
        row = start_row

        # Tier header
        ws.row_dimensions[row].height = 20
        hc = ws.cell(row=row, column=1, value=tier_label)
        hc.fill = _FILL_HEADER
        hc.font = _FONT_HEADER
        hc.border = _THIN_BORDER
        ac = ws.cell(row=row, column=2, value="Amount")
        ac.fill = _FILL_HEADER
        ac.font = _FONT_HEADER
        ac.alignment = Alignment(horizontal="center")
        ac.border = _THIN_BORDER
        row += 1

        for i, (label, val) in enumerate(items):
            ws.row_dimensions[row].height = 17
            lc = ws.cell(row=row, column=1, value=label)
            vc = ws.cell(row=row, column=2, value=val)
            alt = _FILL_ALT_ROW if i % 2 == 1 else PatternFill()
            lc.fill = alt
            vc.fill = alt
            lc.font = _FONT_BODY
            vc.font = _FONT_BODY
            lc.border = _THIN_BORDER
            vc.border = _THIN_BORDER
            vc.number_format = _FMT_AMOUNT
            vc.alignment = Alignment(horizontal="right")
            row += 1

        # Subtotal row
        ws.row_dimensions[row].height = 18
        _apply_subtotal_row(ws, row, [f"Subtotal — {tier_label.split(':')[0]}", subtotal])
        ws.cell(row=row, column=2).number_format = _FMT_AMOUNT
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="right")
        row += 1

        return row

    # ------------------------------------------------------------------
    # Sheet 6: Forensic Summary
    # ------------------------------------------------------------------

    def _build_forensic_summary(self, wb: Workbook, sys: "ColettiOS"):
        ws = wb.create_sheet("Forensic Summary")
        ws.sheet_view.showGridLines = False

        _set_col_widths(ws, [
            (1, 38), (2, 22),
        ])

        row = 1
        ws.row_dimensions[row].height = 24
        t = ws.cell(row=row, column=1, value="FORENSIC ANALYSIS SUMMARY")
        t.font = _FONT_SUBTITLE
        row += 2

        disparity = sys.income_disparity

        # ── Income disparity block ────────────────────────────────────
        ws.row_dimensions[row].height = 20
        _apply_header_row(ws, row, ["Metric", "Value"])
        row += 1

        key_metrics = [
            ("Sworn Monthly Net Income", disparity.sworn_monthly_net),
            ("Verified Monthly Net Income", disparity.verified_monthly_net),
            ("Monthly Understatement Delta", disparity.monthly_understatement()),
            ("Concealment %", disparity.understatement_pct() / 100),
            ("Tracking Period (months)", disparity.tracking_months),
            ("Cumulative Understatement", disparity.cumulative_understatement()),
            ("Sequestered Hard Assets", disparity.sequestered_hard_assets),
            ("Total Concealed Value", disparity.total_concealed_value()),
        ]

        for i, (label, val) in enumerate(key_metrics):
            ws.row_dimensions[row].height = 17
            alt = _FILL_ALT_ROW if i % 2 == 1 else PatternFill()
            lc = ws.cell(row=row, column=1, value=label)
            lc.fill = alt
            lc.font = _FONT_BODY
            lc.border = _THIN_BORDER
            vc = ws.cell(row=row, column=2, value=val)
            vc.fill = alt
            vc.border = _THIN_BORDER
            vc.alignment = Alignment(horizontal="right")
            if isinstance(val, float) and "%" not in label:
                vc.number_format = _FMT_AMOUNT
                vc.font = _FONT_BODY
            elif "%" in label:
                vc.number_format = '0.00%'
                vc.font = _FONT_BODY
            else:
                vc.font = _FONT_BODY
            row += 1

        # Total shielded capital highlight
        ws.row_dimensions[row].height = 20
        ws.cell(row=row, column=1, value="TOTAL SHIELDED CAPITAL").fill = _FILL_GRAND_TOTAL
        ws.cell(row=row, column=1).font = _FONT_GRAND_TOTAL
        ws.cell(row=row, column=1).border = _THIN_BORDER
        tv = ws.cell(row=row, column=2, value=disparity.total_concealed_value())
        tv.fill = _FILL_GRAND_TOTAL
        tv.font = _FONT_GRAND_TOTAL
        tv.number_format = _FMT_AMOUNT
        tv.alignment = Alignment(horizontal="right")
        tv.border = _THIN_BORDER
        row += 2

        # ── Case valuation quick reference ────────────────────────────
        ws.row_dimensions[row].height = 20
        st = ws.cell(row=row, column=1, value="CASE VALUATION QUICK REFERENCE")
        st.font = _FONT_SUBTITLE
        row += 1

        _apply_header_row(ws, row, ["Tier", "Amount"])
        row += 1

        cv = sys.case_valuation
        val_rows = [
            ("Tier 1 — Motion Relief", cv.tier1.subtotal),
            ("Tier 2 — Trial Damages", cv.tier2.subtotal),
            ("Tier 3 — Punitive", cv.tier3.total),
            ("TOTAL (Capped)", cv.total_capped),
        ]
        for i, (label, val) in enumerate(val_rows):
            ws.row_dimensions[row].height = 17
            is_total = label.startswith("TOTAL")
            fill = _FILL_GRAND_TOTAL if is_total else (
                _FILL_ALT_ROW if i % 2 == 1 else PatternFill()
            )
            font = _FONT_GRAND_TOTAL if is_total else _FONT_BODY
            lc = ws.cell(row=row, column=1, value=label)
            lc.fill = fill
            lc.font = font
            lc.border = _THIN_BORDER
            vc = ws.cell(row=row, column=2, value=val)
            vc.fill = fill
            vc.font = font
            vc.number_format = _FMT_AMOUNT
            vc.alignment = Alignment(horizontal="right")
            vc.border = _THIN_BORDER
            row += 1

        row += 2

        # ── Forensic engine data (if available) ───────────────────────
        try:
            from forensic_engine import ForensicEngine
            engine = ForensicEngine()
            va = engine.variance_analysis
            ci = engine.cumulative_impact

            ws.row_dimensions[row].height = 20
            fe_title = ws.cell(row=row, column=1, value="FORENSIC ENGINE ANALYSIS")
            fe_title.font = _FONT_SUBTITLE
            row += 1

            _apply_header_row(ws, row, ["Forensic Metric", "Value"])
            row += 1

            fe_rows = [
                ("Sworn Monthly Net", va.get("sworn_monthly_net", 0)),
                ("Verified Monthly Net", va.get("actual_monthly_net", 0)),
                ("Monthly Concealment Delta", va.get("monthly_concealment_delta", 0)),
                ("Annual Concealment", va.get("annual_concealment_delta", 0)),
                ("Total Concealed Income", ci.get("total_concealed_income", 0)),
                ("Total Support Arrearage", ci.get("total_support_arrearage", 0)),
                ("Concealed Assets Value", ci.get("concealed_assets_value", 0)),
                ("Total Shielded Capital", ci.get("total_shielded_capital", 0)),
            ]
            for i, (label, val) in enumerate(fe_rows):
                ws.row_dimensions[row].height = 17
                alt = _FILL_ALT_ROW if i % 2 == 1 else PatternFill()
                lc = ws.cell(row=row, column=1, value=label)
                lc.fill = alt
                lc.font = _FONT_BODY
                lc.border = _THIN_BORDER
                vc = ws.cell(row=row, column=2, value=val)
                vc.fill = alt
                vc.font = _FONT_BODY
                vc.number_format = _FMT_AMOUNT
                vc.alignment = Alignment(horizontal="right")
                vc.border = _THIN_BORDER
                row += 1
        except Exception:
            pass  # ForensicEngine is optional enrichment


# ---------------------------------------------------------------------------
# Utility to safely retrieve ColettiOS version
# ---------------------------------------------------------------------------

def ColettiOS_VERSION_FROM_SYS(sys_instance) -> str:
    try:
        return f"ColettiOS v{sys_instance.VERSION}"
    except AttributeError:
        return "ColettiOS v2.5.5"
